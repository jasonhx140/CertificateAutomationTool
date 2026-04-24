import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import datetime
import sys

# ===================== 打包兼容修复 =====================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ===================== 配置 =====================
CONFIG_FILE = "config.json"
MERGE_COL_NAME = "销售凭证"

# 必须删除的物料编码
REMOVE_MATERIALS = {
    "590000013", "590000014", "590000015", "590000027", "590000028",
    "590000029", "590000030", "590000003", "590000004", "590000005",
    "590000006", "590000007", "590000008", "590000009", "590000010",
    "590000011", "590000012", "590000032", "590000033", "590000031",
    "590000000"
}

ORG_CONFIG = {
    "维通利华北京销售组织": {
        "test_sheet": "北京",
        "license": "SCXK（京）2021-0011"
    },
    "维通利华湖北销售组织": {
        "test_sheet": "湖北",
        "license": "SCXK（鄂）2022-0030"
    }
}

# ===================== 配置记忆 =====================
def load_config():
    try:
        if os.path.exists(resource_path(CONFIG_FILE)):
            with open(resource_path(CONFIG_FILE), "r", encoding="utf-8") as f:
                return json.load(f)
    except:
        pass
    return {}

def save_config(export_path, cert_path, test_path, selected_org_list):
    cfg = {
        "export": export_path,
        "cert": cert_path,
        "test": test_path,
        "selected_org_list": selected_org_list
    }
    try:
        with open(resource_path(CONFIG_FILE), "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except:
        pass

# ===================== 输出文件夹 =====================
def get_output_folder():
    base_folder = "已生成的合格证"
    if not os.path.exists(base_folder):
        os.makedirs(base_folder)
    return base_folder

# ===================== 文件名：日期+版本号 =====================
def get_filename(prefix):
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    base_folder = get_output_folder()
    filename = f"{prefix}_{today}.xlsx"
    full_path = os.path.join(base_folder, filename)
    version = 2
    while os.path.exists(full_path):
        filename = f"{prefix}_{today}_V{version}.xlsx"
        full_path = os.path.join(base_folder, filename)
        version += 1
    return full_path

# ===================== 自动列宽 + 筛选 =====================
def save_pretty_excel(df, filename, text_col_name):
    # 延迟导入：仅在需要时加载
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "合格证"
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            # 关键修改：如果值是NaN则写入空字符串
            cell_value = value if pd.notna(value) else ""
            ws.cell(row=r_idx, column=c_idx, value=cell_value)

    font_normal = Font(name="微软雅黑", size=9)
    font_header = Font(name="微软雅黑", size=9, bold=True)
    fill_header = PatternFill(start_color="4472C9", end_color="4472C9", fill_type="solid")
    align_center = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = font_normal
            cell.alignment = align_center

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(filename)

# ===================== 核心处理 =====================
def process_all(export_path, test_path, cert_template_path, log_widget, selected_org_list):
    # 延迟导入：仅在点击"生成"时加载重型库，加速 GUI 启动
    import pandas as pd

    def log(msg):
        now = datetime.datetime.now().strftime("%H:%M:%S")
        log_widget.insert(tk.END, f"[{now}] {msg}\n")
        log_widget.see(tk.END)
        log_widget.update()

    try:
        log("开始读取数据源...")
        export_df = pd.read_excel(export_path, dtype=str)
        cert_template = pd.read_excel(cert_template_path, dtype=str)
        log(f"原始数据：{len(export_df)} 行")

        if "物料" in export_df.columns:
            export_df["物料"] = export_df["物料"].astype(str).str.strip()
            before = len(export_df)
            export_df = export_df[~export_df["物料"].isin(REMOVE_MATERIALS)]
            log(f"✅ 已按表头【物料】删除无效数据：{before - len(export_df)} 行")

        if MERGE_COL_NAME in export_df.columns:
            export_df = export_df[~export_df[MERGE_COL_NAME].str.startswith("2", na=False)]

        condition_exclude = (export_df["单位"] == "EA") | export_df["拒绝原因描述"].notna()
        export_df = export_df[~condition_exclude]

        for col in ["最小体重", "最大体重"]:
            if col in export_df.columns:
                export_df[col] = export_df[col].astype(str).str.strip()
                export_df[col] = export_df[col].replace({"0": "", "0.0": ""})
                log(f"✅ {col} 为0时已置空")

        if "性别" in export_df.columns:
            export_df["性别"] = export_df["性别"].replace({"F": "雌", "M": "雄"})
            log("✅ 性别已转换：F→雌，M→雄")

        if "品系" in export_df.columns:
            export_df["品系"] = export_df["品系"].replace({
                "APOE-KO": "APOE",
                "C57BL/6JNifdc Aged": "C57BL/6JNifdc"
            })
            log("✅ 品系已标准化")

        date_cols = [c for c in export_df.columns if "日期" in c or "时间" in c]
        for col in date_cols:
            export_df[col] = pd.to_datetime(export_df[col], errors="coerce").dt.strftime("%Y-%m-%d")

        export_df["合并凭证"] = export_df["销售凭证"].str.strip() + export_df["销售订单行号"].str.strip()
        output_dir = get_output_folder()
        log(f"输出文件夹：{output_dir}")

        for selected_org in selected_org_list:
            if selected_org not in ORG_CONFIG:
                log(f"⚠️ 公司 {selected_org} 未配置，跳过")
                continue
            log(f"\n==================== 开始处理：{selected_org} ====================")
            org_data = export_df[export_df["销售组织描述"] == selected_org].copy()
            log(f"当前公司有效数据：{len(org_data)} 行")
            if len(org_data) == 0:
                log(f"⚠️ {selected_org} 无有效数据，跳过")
                continue
            cfg = ORG_CONFIG[selected_org]
            test_df = pd.read_excel(test_path, sheet_name=cfg["test_sheet"], dtype=str)

            if "品" in test_df.columns:
                key_col = "品系"
            elif "SAP系统品系名称" in test_df.columns:
                key_col = "SAP系统品系名称"
            elif "SAP对应名称" in test_df.columns:
                key_col = "SAP对应名称"
            else:
                raise Exception("未找到品系相关列")

            test_df["检测日期"] = pd.to_datetime(test_df["检测日期"], errors="coerce").dt.strftime("%Y-%m-%d")
            test_df = test_df.drop_duplicates(subset=[key_col], keep="last")
            test_df["key_col_lower"] = test_df[key_col].str.strip().str.lower()
            date_map = dict(zip(test_df["key_col_lower"], test_df["检测日期"]))
            if "品系" in org_data.columns and "最后一次检测日期" in cert_template.columns:
                org_data["品系_lower"] = org_data["品系"].str.strip().str.lower()
                org_data["最后一次检测日期"] = org_data["品系_lower"].map(date_map)
                org_data = org_data.drop(columns=["品系_lower"])

            final_df = pd.DataFrame(columns=cert_template.columns)
            for col in final_df.columns:
                if col in org_data.columns and col != "备注":
                    final_df[col] = org_data[col]
            final_df[MERGE_COL_NAME] = org_data["合并凭证"]
            if "备注" in final_df.columns:
                final_df["备注"] = ""

            # 关键修改：生产许可证号赋值逻辑 - 仅当列存在时赋值，空值保留
            fixed_fields = {
                "生产许可证号": cfg["license"],
                "用途": "科学研究",
                "质检单位": "北京维通利华实验动物技术有限公司",
                "质量负责人": "韩雪",
                "质量等级": "SPF级"
            }
            for k, v in fixed_fields.items():
                if k in final_df.columns:
                    # 仅赋值非空值，若v为空则不修改（保留原空值）
                    if pd.notna(v) and v != "":
                        final_df[k] = v
                    # 若v为空，保持列原有空值，不做赋值

            # E列格式修正逻辑保持不变，但处理空值
            if len(final_df.columns) >= 5:
                e_col_name = final_df.columns[4]
                if e_col_name not in ["", None]:
                    # 先替换空值为""，再处理格式
                    final_df[e_col_name] = final_df[e_col_name].fillna("").astype(str)
                    final_df[e_col_name] = final_df[e_col_name].str.replace(" ", "")
                    final_df[e_col_name] = final_df[e_col_name].str.replace("(", "（").str.replace(")", "）")
                    log(f"✅ {e_col_name} 列格式已修正")

            out_file = get_filename(f"合格证_{cfg['test_sheet']}")
            save_pretty_excel(final_df, out_file, MERGE_COL_NAME)
            log(f"✅ 已生成：{os.path.basename(out_file)}")

        log("\n==================================================")
        log("✅ 所有选中公司处理完成！")
        if messagebox.askyesno("完成", f"全部文件已保存至：\n{output_dir}\n是否打开文件夹？"):
            os.startfile(output_dir)
    except Exception as e:
        log(f"❌ 错误：{str(e)}")
        messagebox.showerror("处理失败", str(e))

# ===================== GUI界面 =====================
def main_gui():
    root = tk.Tk()
    root.title("实验动物合格证自动生成工具")

    # ✅ 唯一正确修改：从底部增加60，不越界、不闪退
    root.geometry("760x585")
    root.resizable(False, False)

    PRIMARY = "#2C70C9"
    SECONDARY = "#F5F7FA"
    ACCENT = "#3E8D66"
    FG = "#FFFFFF"

    top_frame = tk.Frame(root, bg=PRIMARY, height=80)
    top_frame.pack(fill=tk.X)
    tk.Label(top_frame, text="实验动物合格证自动生成工具", font=("微软雅黑", 18, "bold"),
             bg=PRIMARY, fg=FG).place(x=30, y=15)
    tk.Label(top_frame, text="文件记忆｜一键生成｜公司筛选", font=("微软雅黑", 10),
             bg=PRIMARY, fg="#D0E0F0").place(x=32, y=50)

    content_frame = tk.Frame(root, bg=SECONDARY, padx=20, pady=10)
    content_frame.pack(fill=tk.BOTH, expand=True)
    content_frame.pack_propagate(False)

    cfg = load_config()
    paths = {
        "export": cfg.get("export", ""),
        "cert": cfg.get("cert", ""),
        "test": cfg.get("test", "")
    }
    default_selected = cfg.get("selected_org_list", list(ORG_CONFIG.keys()))

    PANEL_WIDTH = 700
    SPACE = 8
    Y = 5

    # 公司选择框
    ORG_H = 85
    org_frame = tk.LabelFrame(content_frame, text="公司选择（可多选）", font=("微软雅黑", 10, "bold"), bg=SECONDARY)
    org_frame.place(x=0, y=Y, width=PANEL_WIDTH, height=ORG_H)
    Y += ORG_H + SPACE

    tk.Label(org_frame, text="请选择需要生成合格证的公司：", font=("微软雅黑", 9), bg=SECONDARY).place(x=15, y=15)
    all_orgs = list(ORG_CONFIG.keys())
    org_vars = {}
    start_y = 18
    step_y = 26
    for i, org in enumerate(all_orgs):
        var = tk.BooleanVar(value=org in default_selected)
        org_vars[org] = var
        tk.Checkbutton(org_frame, text=org, variable=var, font=("微软雅黑", 9), bg=SECONDARY) \
            .place(x=230, y=start_y + i * step_y, anchor="w")

    # 文件选择框
    FILE_H = 125
    file_frame = tk.LabelFrame(content_frame, text="文件选择", font=("微软雅黑", 10, "bold"), bg=SECONDARY)
    file_frame.place(x=0, y=Y, width=PANEL_WIDTH, height=FILE_H)
    Y += FILE_H + SPACE

    rows = [("1. export 源数据文件", "export", 10), ("2. 合格证模板", "cert", 40), ("3. 检测日期报告文件", "test", 70)]
    entrys = {}
    for label, key, y in rows:
        tk.Label(file_frame, text=label, font=("微软雅黑", 9), bg=SECONDARY).place(x=15, y=y)
        ent = tk.Entry(file_frame, font=("微软雅黑", 9), state="readonly", bg="white")
        ent.place(x=180, y=y - 2, width=380, height=22)
        entrys[key] = ent
        if paths[key]:
            ent.config(state=tk.NORMAL)
            ent.insert(0, paths[key])
            ent.config(state="readonly")

        def choose(k=key, e=ent):
            f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
            if f:
                e.config(state=tk.NORMAL)
                e.delete(0, tk.END)
                e.insert(0, f)
                e.config(state="readonly")
                paths[k] = f
                selected = [o for o, v in org_vars.items() if v.get()]
                save_config(paths["export"], paths["cert"], paths["test"], selected)

        tk.Button(file_frame, text="选择文件", width=11, command=choose).place(x=570, y=y - 3)

    # 日志框
    LOG_H = 120
    log_frame = tk.LabelFrame(content_frame, text="运行日志", font=("微软雅黑", 10, "bold"), bg=SECONDARY)
    log_frame.place(x=0, y=Y, width=PANEL_WIDTH, height=LOG_H)
    Y += LOG_H + SPACE

    log_txt = tk.Text(log_frame, font=("Consolas", 9), bg="white", relief=tk.FLAT)
    log_txt.place(x=8, y=6, width=PANEL_WIDTH - 25, height=LOG_H - 35)
    scr = ttk.Scrollbar(log_frame, command=log_txt.yview)
    scr.place(x=PANEL_WIDTH - 15, y=6, width=12, height=LOG_H - 35)
    log_txt.config(yscrollcommand=scr.set)

    # 操作框
    BTN_H = 95
    btn_frame = tk.LabelFrame(content_frame, text="操作", font=("微软雅黑", 10, "bold"), bg=SECONDARY)
    btn_frame.place(x=0, y=Y, width=PANEL_WIDTH, height=BTN_H)

    def run():
        try:
            if not all(paths.values()):
                messagebox.showwarning("提示", "请选择全部3个文件")
                return
            selected_org_list = [o for o, v in org_vars.items() if v.get()]
            if not selected_org_list:
                messagebox.showwarning("提示", "请至少勾选一个公司")
                return
            save_config(paths["export"], paths["cert"], paths["test"], selected_org_list)
            process_all(paths["export"], paths["test"], paths["cert"], log_txt, selected_org_list)
        except Exception as e:
            messagebox.showerror("异常", f"运行出错：{str(e)}")

    tk.Button(btn_frame, text="🚀 一键生成合格证", font=("微软雅黑", 12, "bold"),
              bg=ACCENT, fg=FG, relief=tk.FLAT, command=run) \
        .place(relx=0.5, rely=0.5, anchor="center", width=PANEL_WIDTH - 40, height=45)

    # 底部
    bottom = tk.Frame(root, bg=PRIMARY, height=24)
    bottom.pack(side=tk.BOTTOM, fill=tk.X)
    tk.Label(bottom, text="© 内部专用工具", font=("微软雅黑", 8), bg=PRIMARY, fg="#D0E0F0").pack(anchor="center")

    root.mainloop()

if __name__ == "__main__":
    main_gui()